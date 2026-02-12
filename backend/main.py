"""
Hydrological Modeling – Discharge API (read-only).
Serves discharge stations and time series from discharge.xlsx. No watershed or DEM.
"""
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

_THIS_DIR = Path(__file__).resolve().parent
if str(_THIS_DIR) not in sys.path:
    sys.path.insert(0, str(_THIS_DIR))

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

app = FastAPI(title="Hydrological Modeling API", description="Discharge stations and time series")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET"],
    allow_headers=["*"],
)

_resolve_dir = Path(__file__).resolve().parent.parent
DISCHARGE_DIR = Path(os.environ.get("SWAT_VIEWER_DIR", str(_resolve_dir)))
DISCHARGE_FILE = DISCHARGE_DIR / "discharge.xlsx"
if not DISCHARGE_FILE.exists():
    DISCHARGE_FILE = DISCHARGE_DIR / "discharge.xls"
if not DISCHARGE_FILE.exists():
    DISCHARGE_FILE = _resolve_dir / "discharge.xlsx"
if not DISCHARGE_FILE.exists():
    DISCHARGE_FILE = _resolve_dir / "discharge.xls"

_discharge_cache = {"mtime": None, "stations": None, "series": None, "station_ids": None}


def _norm_station_id(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def load_discharge_data():
    if not DISCHARGE_FILE.exists():
        return None, None
    mtime = DISCHARGE_FILE.stat().st_mtime
    if _discharge_cache["mtime"] == mtime and _discharge_cache["series"] is not None:
        return _discharge_cache["station_ids"], _discharge_cache["series"]
    wb = load_workbook(DISCHARGE_FILE, read_only=True, data_only=True)
    data_sheet = None
    for name in ("data", "Merged", "Data", "merged"):
        if name in wb.sheetnames:
            data_sheet = wb[name]
            break
    if data_sheet is None:
        data_sheet = wb[wb.sheetnames[0]]
    rows = list(data_sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None, None
    header = rows[0]
    station_ids = [_norm_station_id(h) for h in header[1:]]
    series = {}
    for st_id in station_ids:
        if st_id:
            series[st_id] = []
    for row in rows[1:]:
        if not row:
            continue
        cell0 = row[0]
        if cell0 is None:
            continue
        if isinstance(cell0, datetime):
            date_str = cell0.strftime("%Y-%m-%d")
        else:
            date_str = str(cell0).strip()[:10]
        for i, st_id in enumerate(station_ids):
            if not st_id:
                continue
            col_idx = i + 1
            if col_idx >= len(row):
                continue
            try:
                v = row[col_idx]
                val = float(v) if v is not None and str(v).strip() != "" else None
            except (ValueError, TypeError):
                val = None
            series[st_id].append({"date": date_str, "value": val})
    _discharge_cache["mtime"] = mtime
    _discharge_cache["station_ids"] = station_ids
    _discharge_cache["series"] = series
    return station_ids, series


def load_discharge_coordinates():
    if not DISCHARGE_FILE.exists():
        return []
    mtime = DISCHARGE_FILE.stat().st_mtime
    if _discharge_cache["mtime"] == mtime and _discharge_cache["stations"] is not None:
        return _discharge_cache["stations"]
    wb = load_workbook(DISCHARGE_FILE, read_only=True, data_only=True)
    coord_sheet = None
    for name in ("coordinates", "Coordinates", "COORDINATES"):
        if name in wb.sheetnames:
            coord_sheet = wb[name]
            break
    if coord_sheet is None:
        wb.close()
        station_ids, _ = load_discharge_data()
        if station_ids:
            out = [{"id": sid, "name": sid, "staname": None, "lat": None, "lon": None} for sid in station_ids if sid]
            _discharge_cache["stations"] = out
            return out
        return []
    rows = list(coord_sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    raw_headers = rows[0]

    def norm(s):
        return (str(s).strip().lower().replace(" ", "") if s is not None else "")

    staid_idx = staname_idx = lat_idx = lon_idx = None
    for i, c in enumerate(raw_headers):
        h = norm(c)
        if h == "staid":
            staid_idx = i
        elif h in ("staname", "stationname") or "staname" in h or (c and "name" in str(c).lower() and "sta" in str(c).lower()):
            staname_idx = i
        elif h in ("lat", "latitude", "y"):
            lat_idx = i
        elif h in ("lon", "long", "longitude", "lng", "x"):
            lon_idx = i
    if staid_idx is None:
        return []
    out = []
    for row in rows[1:]:
        if staid_idx >= len(row):
            continue
        staid = _norm_station_id(row[staid_idx])
        if not staid:
            continue
        staname = None
        if staname_idx is not None and staname_idx < len(row) and row[staname_idx]:
            staname = str(row[staname_idx]).strip()
        name = staname if staname else staid
        lat = None
        lon = None
        if lat_idx is not None and lat_idx < len(row) and row[lat_idx] is not None:
            try:
                lat = float(row[lat_idx])
            except (ValueError, TypeError):
                pass
        if lon_idx is not None and lon_idx < len(row) and row[lon_idx] is not None:
            try:
                lon = float(row[lon_idx])
            except (ValueError, TypeError):
                pass
        out.append({"id": staid, "name": name, "staname": staname, "lat": lat, "lon": lon})
    load_discharge_data()
    _discharge_cache["stations"] = out
    return out


@app.get("/")
def root():
    return {
        "name": "Hydrological Modeling API",
        "version": "2.0",
        "endpoints": ["/api/discharge/stations", "/api/discharge/station/{station_id}"],
    }


@app.get("/api/discharge/stations")
def get_discharge_stations():
    stations = load_discharge_coordinates()
    features = []
    for s in stations:
        sid = str(s["id"]) if s.get("id") is not None else ""
        name = s.get("name") or sid
        props = {"name": name, "staname": s.get("staname"), "id": sid}
        if s["lat"] is not None and s["lon"] is not None:
            features.append({
                "type": "Feature",
                "id": sid,
                "properties": props,
                "geometry": {"type": "Point", "coordinates": [s["lon"], s["lat"]]},
            })
        else:
            features.append({
                "type": "Feature",
                "id": sid,
                "properties": props,
                "geometry": None,
            })
    return {"type": "FeatureCollection", "features": features}


@app.get("/api/discharge/station/{station_id}")
def get_discharge_station(
    station_id: str,
    limit: Optional[int] = Query(50000, description="Max number of points to return"),
):
    _, series = load_discharge_data()
    if series is None:
        raise HTTPException(status_code=404, detail="Discharge data not loaded")
    sid_norm = _norm_station_id(station_id)
    if sid_norm not in series:
        for k in series:
            if _norm_station_id(k) == sid_norm or k == station_id:
                sid_norm = k
                break
        else:
            raise HTTPException(status_code=404, detail=f"Station '{station_id}' not found")
    data = series[sid_norm]
    if limit and len(data) > limit:
        data = data[-limit:]
    stations = load_discharge_coordinates()
    display_name = next((s["name"] for s in stations if _norm_station_id(s["id"]) == sid_norm), sid_norm)
    return {"id": sid_norm, "name": display_name, "discharge": data}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
