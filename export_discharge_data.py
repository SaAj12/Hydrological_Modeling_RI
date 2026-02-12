"""
Export discharge.xlsx to JSON for static web deployment (e.g. GitHub Pages).
Run from project root: python scripts/export_discharge_data.py
Output: frontend/data/discharge_data.json
"""
import json
import sys
from datetime import datetime
from pathlib import Path

# Project root = parent of scripts/
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DISCHARGE_FILE = PROJECT_ROOT / "discharge.xlsx"
OUTPUT_FILE = PROJECT_ROOT / "frontend" / "data" / "discharge_data.json"


def main():
    if not DISCHARGE_FILE.exists():
        print(f"Error: {DISCHARGE_FILE} not found")
        sys.exit(1)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(DISCHARGE_FILE, read_only=True, data_only=True)

    # Load coordinates
    stations = []
    coord_sheet = None
    for name in ("coordinates", "Coordinates", "COORDINATES"):
        if name in wb.sheetnames:
            coord_sheet = wb[name]
            break
    if coord_sheet:
        rows = list(coord_sheet.iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip().lower().replace(" ", "") if h else "" for h in rows[0]]
            staid_idx = next((i for i, h in enumerate(headers) if h == "staid"), None)
            staname_idx = next((i for i, h in enumerate(headers) if h in ("staname", "stationname", "stations_names", "station_names") or "staname" in h or ("sta" in h and "name" in h)), None)
            lat_idx = next((i for i, h in enumerate(headers) if h in ("lat", "latitude", "y")), None)
            lon_idx = next((i for i, h in enumerate(headers) if h in ("lon", "long", "longitude", "lng", "x")), None)
            if staid_idx is not None:
                for row in rows[1:]:
                    if staid_idx >= len(row) or not row[staid_idx]:
                        continue
                    staid = str(row[staid_idx]).strip()
                    staname = None
                    if staname_idx is not None and staname_idx < len(row) and row[staname_idx]:
                        staname = str(row[staname_idx]).strip()
                    lat = float(row[lat_idx]) if lat_idx is not None and lat_idx < len(row) and row[lat_idx] is not None else None
                    lon = float(row[lon_idx]) if lon_idx is not None and lon_idx < len(row) and row[lon_idx] is not None else None
                    stations.append({
                        "id": staid,
                        "name": staname or staid,
                        "lat": lat,
                        "lon": lon
                    })
    wb.close()

    # Load data sheet
    wb = load_workbook(DISCHARGE_FILE, read_only=True, data_only=True)
    data_sheet = None
    for name in ("data", "Merged", "Data", "merged"):
        if name in wb.sheetnames:
            data_sheet = wb[name]
            break
    if data_sheet is None:
        data_sheet = wb[wb.sheetnames[0]]
    rows = list(data_sheet.iter_rows(values_only=True))
    wb.close()

    series = {}
    if rows:
        header = rows[0]
        station_ids = [str(h).strip() for h in header[1:] if h]
        for sid in station_ids:
            if sid:
                series[sid] = []
        for row in rows[1:]:
            if not row:
                continue
            cell0 = row[0]
            if cell0 is None:
                continue
            if isinstance(cell0, datetime):
                date_str = cell0.strftime("%Y-%m-%d")
            else:
                date_str = str(cell0).strip()[:10]
            for i, sid in enumerate(station_ids):
                if not sid:
                    continue
                col_idx = i + 1
                if col_idx >= len(row):
                    continue
                try:
                    v = row[col_idx]
                    val = float(v) if v is not None and str(v).strip() != "" else None
                except (ValueError, TypeError):
                    val = None
                series[sid].append({"date": date_str, "value": val})

    # Add stations that are in series but not in coordinates
    for sid in series:
        if not any(s["id"] == sid for s in stations):
            stations.append({"id": sid, "name": sid, "lat": None, "lon": None})

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    out = {"stations": stations, "series": series}
    with open(OUTPUT_FILE, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Exported {len(stations)} stations, {len(series)} time series to {OUTPUT_FILE}")

    # Also update docs/ if it exists (for GitHub Pages)
    docs_data = PROJECT_ROOT / "docs" / "data" / "discharge_data.json"
    if (PROJECT_ROOT / "docs").exists():
        docs_data.parent.mkdir(parents=True, exist_ok=True)
        with open(docs_data, "w") as f:
            json.dump(out, f, indent=2)
        print(f"Also updated {docs_data} for GitHub Pages")


if __name__ == "__main__":
    main()
