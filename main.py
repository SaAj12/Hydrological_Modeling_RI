"""
SWAT+ Climate & Flow Viewer – Backend API (read-only).
Serves grid locations and time series (precipitation, temperature, etc.) for the map frontend.
Discharge data from discharge.xlsx: "data" sheet (or "Merged") + "coordinates" sheet (STAID, lat, lon).
"""
import csv
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

# Ensure we load from this file's directory (avoid wrong main.py from elsewhere)
_THIS_DIR = Path(__file__).resolve().parent
if str(_THIS_DIR) not in sys.path:
    sys.path.insert(0, str(_THIS_DIR))

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from openpyxl import load_workbook

app = FastAPI(title="SWAT+ Climate Viewer API", description="Read-only API for grid climate data")

# Allow frontend on GitHub Pages or any origin
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET"],
    allow_headers=["*"],
)

# Data directory: domain folder with pcp/, tmp/, grids_domain.csv
DATA_DIR = Path(os.environ.get("SWAT_DATA_DIR", ""))
if not DATA_DIR or not DATA_DIR.is_dir():
    DATA_DIR = Path(__file__).resolve().parent.parent / "climate_data" / "domain_41_42.5_70.5_72.5"
if not DATA_DIR.is_dir():
    DATA_DIR = Path(__file__).resolve().parent / "data" / "domain_41_42.5_70.5_72.5"
if not DATA_DIR.is_dir():
    # Local dev: sibling repo swat_climate_download
    DATA_DIR = Path(__file__).resolve().parent.parent.parent / "swat_climate_download" / "climate_data" / "domain_41_42.5_70.5_72.5"

# Discharge Excel: project root discharge.xlsx or discharge.xls (use .xlsx path; openpyxl needs .xlsx)
_resolve_dir = Path(__file__).resolve().parent.parent
DISCHARGE_DIR = Path(os.environ.get("SWAT_VIEWER_DIR", str(_resolve_dir)))
DISCHARGE_FILE = DISCHARGE_DIR / "discharge.xlsx"
if not DISCHARGE_FILE.exists():
    DISCHARGE_FILE = DISCHARGE_DIR / "discharge.xls"
if not DISCHARGE_FILE.exists():
    DISCHARGE_FILE = _resolve_dir / "discharge.xlsx"
if not DISCHARGE_FILE.exists():
    DISCHARGE_FILE = _resolve_dir / "discharge.xls"

# Maps: watershed shapefile and DEM raster (project folder maps/)
# Try SWAT_VIEWER_DIR/maps, then script dir parent/maps, then current dir maps
def _maps_dir():
    for d in [DISCHARGE_DIR / "maps", _resolve_dir / "maps", Path.cwd() / "maps"]:
        if d.is_dir():
            return d
    return DISCHARGE_DIR / "maps"

MAPS_DIR = _maps_dir()
WATERSHED_SHP = MAPS_DIR / "watershed_WBDHU8.shp"
DEM_DIR = MAPS_DIR / "dem"

# In-memory cache for discharge data (cleared when file changes)
_discharge_cache = {"mtime": None, "stations": None, "series": None, "station_ids": None}


def _norm_station_id(val) -> str:
    """Normalize station ID for matching (e.g. 1108000 -> '1108000')."""
    if val is None:
        return ""
    return str(val).strip()


def load_discharge_data():
    """Load discharge time series from 'data' or 'Merged' sheet. First row = station IDs, first col = date."""
    if not DISCHARGE_FILE.exists():
        return None, None
    mtime = DISCHARGE_FILE.stat().st_mtime
    if _discharge_cache["mtime"] == mtime and _discharge_cache["series"] is not None:
        return _discharge_cache["station_ids"], _discharge_cache["series"]
    wb = load_workbook(DISCHARGE_FILE, read_only=True, data_only=True)
    data_sheet = None
    for name in ("data", "Merged", "Data", "merged"):
        if name in wb.sheetnames:
            data_sheet = wb[name]
            break
    if data_sheet is None:
        data_sheet = wb[wb.sheetnames[0]]
    rows = list(data_sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None, None
    header = rows[0]
    # First column is date; rest are station IDs (location names)
    station_ids = [_norm_station_id(h) for h in header[1:]]
    series = {}  # station_id -> list of {date, value}
    for st_id in station_ids:
        if st_id:
            series[st_id] = []
    date_col_idx = 0
    for row in rows[1:]:
        if not row:
            continue
        cell0 = row[date_col_idx]
        if cell0 is None:
            continue
        if isinstance(cell0, datetime):
            date_str = cell0.strftime("%Y-%m-%d")
        else:
            try:
                date_str = str(cell0).strip()[:10]
            except Exception:
                continue
        for i, st_id in enumerate(station_ids):
            if not st_id:
                continue
            col_idx = i + 1
            if col_idx >= len(row):
                continue
            try:
                v = row[col_idx]
                if v is not None and str(v).strip() != "":
                    val = float(v)
                else:
                    val = None
            except (ValueError, TypeError):
                val = None
            series[st_id].append({"date": date_str, "value": val})
    _discharge_cache["mtime"] = mtime
    _discharge_cache["station_ids"] = station_ids
    _discharge_cache["series"] = series
    return station_ids, series


def load_discharge_coordinates():
    """Load station coordinates from 'coordinates' sheet. STAID column + lat/lon columns."""
    if not DISCHARGE_FILE.exists():
        return []
    mtime = DISCHARGE_FILE.stat().st_mtime
    if _discharge_cache["mtime"] == mtime and _discharge_cache["stations"] is not None:
        return _discharge_cache["stations"]
    wb = load_workbook(DISCHARGE_FILE, read_only=True, data_only=True)
    coord_sheet = None
    for name in ("coordinates", "Coordinates", "COORDINATES"):
        if name in wb.sheetnames:
            coord_sheet = wb[name]
            break
    if coord_sheet is None:
        wb.close()
        station_ids, _ = load_discharge_data()
        # No coordinates sheet: return stations with null coords so frontend can still list them
        if station_ids:
            out = [{"id": sid, "name": sid, "staname": None, "lat": None, "lon": None} for sid in station_ids if sid]
            _discharge_cache["stations"] = out
            return out
        return []
    rows = list(coord_sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    raw_headers = rows[0]

    def norm(s):
        return (str(s).strip().lower().replace(" ", "") if s is not None else "")

    staid_idx = None
    staname_idx = None
    lat_idx = None
    lon_idx = None
    for i, c in enumerate(raw_headers):
        h = norm(c)
        if h == "staid":
            staid_idx = i
        elif h in ("staname", "stationname") or "staname" in h or (c and "name" in str(c).lower() and "sta" in str(c).lower()):
            staname_idx = i
        elif h in ("lat", "latitude", "y"):
            lat_idx = i
        elif h in ("lon", "long", "longitude", "lng", "x"):
            lon_idx = i
    if staid_idx is None:
        return []
    out = []
    for row in rows[1:]:
        if staid_idx >= len(row):
            continue
        staid = _norm_station_id(row[staid_idx])
        if not staid:
            continue
        staname = None
        if staname_idx is not None and staname_idx < len(row):
            val = row[staname_idx]
            if val is not None and str(val).strip():
                staname = str(val).strip()
        name = staname if staname else staid
        lat = None
        lon = None
        if lat_idx is not None and lat_idx < len(row) and row[lat_idx] is not None:
            try:
                lat = float(row[lat_idx])
            except (ValueError, TypeError):
                pass
        if lon_idx is not None and lon_idx < len(row) and row[lon_idx] is not None:
            try:
                lon = float(row[lon_idx])
            except (ValueError, TypeError):
                pass
        out.append({"id": staid, "name": name, "staname": staname, "lat": lat, "lon": lon})
    load_discharge_data()  # ensure cache mtime is set
    _discharge_cache["stations"] = out
    return out


def load_watershed_geojson():
    """Load watershed shapefile and return GeoJSON FeatureCollection."""
    if not WATERSHED_SHP.exists():
        return None
    try:
        import shapefile
    except ImportError:
        return None
    try:
        sf = shapefile.Reader(str(WATERSHED_SHP))
        field_names = [f[0] for f in sf.fields[1:]]
        features = []
        for shape, record in zip(sf.shapes(), sf.records()):
            if hasattr(shape, "__geo_interface__"):
                geom = shape.__geo_interface__
            else:
                geom = _shape_to_geojson_geom(shape)
            if geom is None:
                continue
            props = dict(zip(field_names, record)) if field_names else {}
            features.append({"type": "Feature", "properties": props, "geometry": geom})
        return {"type": "FeatureCollection", "features": features}
    except Exception:
        return None


def _shape_to_geojson_geom(shape):
    """Convert pyshp shape to GeoJSON geometry dict."""
    if shape.shapeType not in (3, 5):  # POLYLINE, POLYGON
        if shape.shapeType == 1:  # POINT
            return {"type": "Point", "coordinates": shape.points[0]}
        return None
    parts = list(shape.parts) + [len(shape.points)]
    rings = [shape.points[parts[i] : parts[i + 1]] for i in range(len(parts) - 1)]
    if shape.shapeType == 5:  # POLYGON
        return {"type": "Polygon", "coordinates": rings}
    return {"type": "MultiLineString", "coordinates": rings}


def load_grids():
    """Load grid list from grids_domain.csv."""
    path = DATA_DIR / "grids_domain.csv"
    if not path.exists():
        return []
    rows = []
    with open(path) as f:
        r = csv.DictReader(f)
        for row in r:
            rows.append({
                "id": row["ID"],
                "name": row["NAME"].strip(),
                "lat": float(row["LAT"]),
                "lon": float(row["LONG"]),
                "elev": float(row["ELEVATION"]),
            })
    return rows


def parse_pcp(path: Path) -> list[dict]:
    """Parse CHIRPS pcp file: first line YYYYMMDD, then one value per line. Returns [{date, value}]."""
    if not path.exists():
        return []
    with open(path) as f:
        lines = f.readlines()
    if not lines:
        return []
    start = lines[0].strip()
    if len(start) != 8:
        return []
    from datetime import datetime, timedelta
    try:
        base = datetime(int(start[:4]), int(start[4:6]), int(start[6:8]))
    except ValueError:
        return []
    out = []
    for i, line in enumerate(lines[1:], start=1):
        try:
            v = float(line.strip())
        except ValueError:
            v = None
        d = base + timedelta(days=i - 1)
        out.append({"date": d.strftime("%Y-%m-%d"), "value": v})
    return out


def parse_tmp(path: Path) -> list[dict]:
    """Parse CHIRTS tmp file: first line YYYYMMDD, then 'tmax,tmin' per line. Returns [{date, tmax, tmin}]."""
    if not path.exists():
        return []
    with open(path) as f:
        lines = f.readlines()
    if not lines:
        return []
    start = lines[0].strip()
    if len(start) != 8:
        return []
    from datetime import datetime, timedelta
    try:
        base = datetime(int(start[:4]), int(start[4:6]), int(start[6:8]))
    except ValueError:
        return []
    out = []
    for i, line in enumerate(lines[1:], start=1):
        parts = line.strip().split(",")
        tmax = float(parts[0]) if len(parts) >= 1 and parts[0] else None
        tmin = float(parts[1]) if len(parts) >= 2 and parts[1] else None
        d = base + timedelta(days=i - 1)
        out.append({"date": d.strftime("%Y-%m-%d"), "tmax": tmax, "tmin": tmin})
    return out


def parse_slr(path: Path) -> list[dict]:
    """Parse SWAT+ .slr: line 4+ are year, jday, value. Returns [{date, value}]."""
    if not path.exists():
        return []
    with open(path) as f:
        lines = f.readlines()
    if len(lines) < 4:
        return []
    from datetime import datetime
    out = []
    for line in lines[3:]:
        parts = line.strip().split()
        if len(parts) < 3:
            continue
        try:
            y, j, v = int(parts[0]), int(parts[1]), float(parts[2])
            d = datetime(y, 1, 1)
            from datetime import timedelta
            d = d + timedelta(days=j - 1)
            out.append({"date": d.strftime("%Y-%m-%d"), "value": v if v != -99 else None})
        except (ValueError, IndexError):
            continue
    return out


@app.get("/")
def root():
    return {
        "name": "SWAT+ Climate Viewer API",
        "version": "1.0",
        "build": "with-api-info-watershed-dem",
        "loaded_from": str(Path(__file__).resolve()),
        "endpoints": [
            "/api/grids",
            "/api/point/{point_id}",
            "/api/search",
            "/api/discharge/stations",
            "/api/discharge/station/{station_id}",
            "/api/watershed",
            "/api/dem/bounds",
            "/api/dem/image",
            "/api/info",
        ],
    }


@app.get("/api/info")
def api_info():
    """Return paths and availability for discharge, watershed, DEM (for debugging)."""
    stations = load_discharge_coordinates()
    first_has_name = bool(stations and (stations[0].get("name") or stations[0].get("staname")))
    return {
        "discharge_file": str(DISCHARGE_FILE),
        "discharge_exists": DISCHARGE_FILE.exists(),
        "discharge_stations_count": len(stations),
        "stations_have_name": first_has_name,
        "maps_dir": str(MAPS_DIR),
        "maps_dir_exists": MAPS_DIR.is_dir(),
        "watershed_shp": str(WATERSHED_SHP),
        "watershed_exists": WATERSHED_SHP.exists(),
        "dem_dir": str(DEM_DIR),
        "dem_dir_exists": DEM_DIR.is_dir(),
        "pyshp_available": _pyshp_available(),
        "rasterio_available": _rasterio_available(),
    }


def _pyshp_available():
    try:
        import shapefile  # noqa: F401
        return True
    except ImportError:
        return False


def _rasterio_available():
    try:
        import rasterio  # noqa: F401
        return True
    except ImportError:
        return False


@app.get("/api/grids")
def get_grids():
    """Return all grid points for the map (GeoJSON-like or simple list)."""
    grids = load_grids()
    return {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "id": g["id"],
                "properties": {"name": g["name"], "elev": g["elev"]},
                "geometry": {"type": "Point", "coordinates": [g["lon"], g["lat"]]},
            }
            for g in grids
        ],
    }


@app.get("/api/point/{point_id}")
def get_point_data(
    point_id: str,
    pcp: bool = Query(True, description="Include precipitation"),
    tmp: bool = Query(True, description="Include temperature"),
    slr: bool = Query(True, description="Include solar radiation"),
    limit: Optional[int] = Query(365 * 2, description="Max number of days to return (default 2 years)"),
):
    """Return time series for one grid point. point_id is the NAME without 'NA_' or the ID (e.g. 425405 or NA_425405)."""
    grids = {g["id"]: g for g in load_grids()}
    name = None
    if point_id in grids:
        name = grids[point_id]["name"]
    else:
        pid = point_id.replace("NA_", "").strip()
        for g in grids.values():
            if g["id"] == pid or g["name"] == point_id or g["name"] == f"NA_{pid}":
                name = g["name"]
                break
    if not name:
        raise HTTPException(status_code=404, detail="Point not found")
    base = name.replace(".txt", "").replace(".slr", "")
    grid_info = next((g for g in load_grids() if g["name"] == name), None)
    out = {"id": point_id, "name": name, "grid": grid_info}

    if pcp:
        pcp_path = DATA_DIR / "pcp" / f"{base}.txt"
        series = parse_pcp(pcp_path)[-limit:] if limit else parse_pcp(pcp_path)
        out["pcp"] = series

    if tmp:
        tmp_path = DATA_DIR / "tmp" / f"{base}.txt"
        series = parse_tmp(tmp_path)[-limit:] if limit else parse_tmp(tmp_path)
        out["tmp"] = series

    if slr:
        slr_path = DATA_DIR / "slr" / f"{base}.slr"
        if slr_path.exists():
            series = parse_slr(slr_path)[-limit:] if limit else parse_slr(slr_path)
            out["slr"] = series
        else:
            out["slr"] = []

    return out


@app.get("/api/search")
def search(
    lat: float = Query(..., description="Latitude"),
    lon: float = Query(..., description="Longitude"),
):
    """Find the nearest grid point to the given coordinates."""
    grids = load_grids()
    if not grids:
        raise HTTPException(status_code=404, detail="No grids loaded")
    best = min(grids, key=lambda g: (g["lat"] - lat) ** 2 + (g["lon"] - lon) ** 2)
    return {
        "lat": lat,
        "lon": lon,
        "nearest": {
            "id": best["id"],
            "name": best["name"],
            "lat": best["lat"],
            "lon": best["lon"],
            "elev": best["elev"],
        },
    }


# ----- Discharge (Excel: data + coordinates sheets) -----


@app.get("/api/discharge/stations")
def get_discharge_stations():
    """Return all discharge stations (GeoJSON). Stations with lat/lon have Point geometry; others have geometry null so frontend can list them."""
    stations = load_discharge_coordinates()
    features = []
    for s in stations:
        sid = str(s["id"]) if s.get("id") is not None else ""
        name = s.get("name") or sid
        props = {"name": name, "staname": s.get("staname"), "id": sid}
        if s["lat"] is not None and s["lon"] is not None:
            features.append({
                "type": "Feature",
                "id": sid,
                "properties": props,
                "geometry": {"type": "Point", "coordinates": [s["lon"], s["lat"]]},
            })
        else:
            features.append({
                "type": "Feature",
                "id": sid,
                "properties": props,
                "geometry": None,
            })
    return {"type": "FeatureCollection", "features": features}


@app.get("/api/discharge/station/{station_id}")
def get_discharge_station(
    station_id: str,
    limit: Optional[int] = Query(365 * 2, description="Max number of points to return"),
):
    """Return discharge time series for one station. station_id must match STAID or first-row name in data sheet."""
    _, series = load_discharge_data()
    if series is None:
        raise HTTPException(status_code=404, detail="Discharge data not loaded")
    sid_norm = _norm_station_id(station_id)
    if sid_norm not in series:
        for k in series:
            if _norm_station_id(k) == sid_norm or k == station_id:
                sid_norm = k
                break
        else:
            raise HTTPException(status_code=404, detail=f"Station '{station_id}' not found")
    data = series[sid_norm]
    if limit and len(data) > limit:
        data = data[-limit:]
    stations = load_discharge_coordinates()
    display_name = next((s["name"] for s in stations if _norm_station_id(s["id"]) == sid_norm), sid_norm)
    return {"id": sid_norm, "name": display_name, "discharge": data}


# ----- Watershed & DEM (maps folder) -----


@app.get("/api/watershed")
def get_watershed():
    """Return watershed polygons as GeoJSON (from maps/watershed_WBDHU8.shp). Returns empty FeatureCollection if not found."""
    geojson = load_watershed_geojson()
    if geojson is None:
        return {
            "type": "FeatureCollection",
            "features": [],
            "message": f"Watershed not found. Looked at: {WATERSHED_SHP}. Set SWAT_VIEWER_DIR if needed.",
        }
    return geojson


@app.get("/api/dem/bounds")
def get_dem_bounds():
    """Return DEM extent as [south, north, west, east] in WGS84. Returns null if not found."""
    bounds = _get_dem_bounds()
    if bounds is None:
        raise HTTPException(
            status_code=404,
            detail=f"DEM not found. Looked at: {DEM_DIR}. Install rasterio/GDAL for DEM support.",
        )
    return bounds


@app.get("/api/dem/image", response_class=Response)
def get_dem_image():
    """Return DEM as PNG image for L.imageOverlay (same bounds as /api/dem/bounds)."""
    result = _get_dem_image()
    if result is None:
        raise HTTPException(
            status_code=404,
            detail=f"DEM image not found. Looked at: {DEM_DIR}.",
        )
    return Response(content=result, media_type="image/png")


def _get_dem_bounds():
    """Read DEM extent with rasterio; return [south, north, west, east] in WGS84."""
    try:
        import rasterio
        from rasterio.warp import transform_bounds
    except ImportError:
        return None
    if not DEM_DIR.is_dir():
        return None
    try:
        with rasterio.open(str(DEM_DIR)) as src:
            b = src.bounds
            try:
                wgs84_bounds = transform_bounds(src.crs, "EPSG:4326", b.left, b.bottom, b.right, b.top)
                return {"south": wgs84_bounds[1], "north": wgs84_bounds[3], "west": wgs84_bounds[0], "east": wgs84_bounds[2]}
            except Exception:
                return {"south": b.bottom, "north": b.top, "west": b.left, "east": b.right}
    except Exception:
        return None


def _get_dem_image():
    """Read DEM with rasterio, render to PNG (grayscale), return bytes."""
    try:
        import io
        import rasterio
        from rasterio.warp import transform_bounds, reproject, Resampling
        from rasterio.crs import CRS
        import numpy as np
    except ImportError:
        return None
    if not DEM_DIR.is_dir():
        return None
    try:
        with rasterio.open(str(DEM_DIR)) as src:
            data = src.read(1)
            nodata = src.nodata
            if nodata is not None:
                data = np.where(data == nodata, np.nan, data)
            vmin, vmax = np.nanmin(data), np.nanmax(data)
            if vmax <= vmin:
                vmin, vmax = vmin - 1, vmax + 1
            data_byte = np.uint8(255 * (data - vmin) / (vmax - vmin))
            data_byte = np.where(np.isnan(data), 0, data_byte)
            from PIL import Image
            img = Image.fromarray(data_byte, mode="L")
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            return buf.getvalue()
    except Exception:
        return None


if __name__ == "__main__":
    import uvicorn
    _main_path = Path(__file__).resolve()
    print("Starting backend from:", _main_path, flush=True)
    print("Endpoints include: /api/info, /api/watershed, /api/dem/bounds", flush=True)
    uvicorn.run(app, host="0.0.0.0", port=8000)
